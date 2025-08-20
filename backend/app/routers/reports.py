from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from typing import List, Optional
from datetime import datetime, timedelta
import tempfile
import os
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import io

from app.database import get_db, Product, StockEntry, StockExit, StockMovement, StockEntryItem
from app.database import StockExitItem
from app.schemas import User, StockReport, PeriodReport
from app.routers.auth import get_current_active_user

router = APIRouter()

@router.get("/stock-summary", response_model=List[StockReport])
def get_stock_summary(
    date_debut: Optional[datetime] = Query(None),
    date_fin: Optional[datetime] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Résumé du stock par produit avec totaux des entrées et sorties"""
    
    # Requête de base pour les produits
    products = db.query(Product).all()
    
    stock_reports = []
    for product in products:
        # Calculer les totaux d'entrées
        entries_query = db.query(
            func.sum(StockEntryItem.qte_kg).label('total_kg'),
            func.sum(StockEntryItem.qte_cartons).label('total_cartons')
        ).join(StockEntry, StockEntryItem.entry_id == StockEntry.id).filter(StockEntryItem.product_id == product.id)
        
        # Calculer les totaux de sorties
        exits_query = db.query(
            func.sum(StockExitItem.qte_kg).label('total_kg'),
            func.sum(StockExitItem.qte_cartons).label('total_cartons')
        ).join(StockExit, StockExitItem.exit_id == StockExit.id).filter(StockExitItem.product_id == product.id)
        
        # Appliquer les filtres de date si fournis
        if date_debut and date_fin:
            entries_query = entries_query.filter(
                and_(StockEntry.date_reception >= date_debut, StockEntry.date_reception <= date_fin)
            )
            exits_query = exits_query.filter(
                and_(StockExit.date_sortie >= date_debut, StockExit.date_sortie <= date_fin)
            )
        elif date_debut:
            entries_query = entries_query.filter(StockEntry.date_reception >= date_debut)
            exits_query = exits_query.filter(StockExit.date_sortie >= date_debut)
        elif date_fin:
            entries_query = entries_query.filter(StockEntry.date_reception <= date_fin)
            exits_query = exits_query.filter(StockExit.date_sortie <= date_fin)
        
        entries_result = entries_query.first()
        exits_result = exits_query.first()
        
        stock_reports.append(StockReport(
            product=product,
            total_entrees_kg=entries_result.total_kg or 0.0,
            total_entrees_cartons=entries_result.total_cartons or 0,
            total_sorties_kg=exits_result.total_kg or 0.0,
            total_sorties_cartons=exits_result.total_cartons or 0,
            stock_actuel_kg=product.stock_actuel_kg,
            stock_actuel_cartons=product.stock_actuel_cartons
        ))
    
    return stock_reports

@router.get("/period-report", response_model=PeriodReport)
def get_period_report(
    date_debut: datetime = Query(...),
    date_fin: datetime = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Rapport de période avec statistiques globales"""
    
    # Nombre total de produits
    total_produits = db.query(Product).count()
    
    # Nombre d'entrées dans la période
    total_entrees = db.query(StockEntry).filter(
        and_(StockEntry.date_reception >= date_debut, StockEntry.date_reception <= date_fin)
    ).count()
    
    # Nombre de sorties dans la période
    total_sorties = db.query(StockExit).filter(
        and_(StockExit.date_sortie >= date_debut, StockExit.date_sortie <= date_fin)
    ).count()
    
    # Valeur du stock actuel (basée sur les prix d'achat)
    valeur_stock = db.query(
        func.sum(Product.stock_actuel_kg * Product.prix_achat)
    ).scalar() or 0.0
    
    return PeriodReport(
        date_debut=date_debut,
        date_fin=date_fin,
        total_produits=total_produits,
        total_entrees=total_entrees,
        total_sorties=total_sorties,
        valeur_stock=valeur_stock
    )

@router.get("/movements")
def get_movements(
    product_id: Optional[int] = Query(None),
    date_debut: Optional[datetime] = Query(None),
    date_fin: Optional[datetime] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Historique des mouvements (tous produits ou filtré par produit)."""
    query = db.query(StockMovement)
    if product_id:
        query = query.filter(StockMovement.product_id == product_id)
    if date_debut and date_fin:
        query = query.filter(and_(StockMovement.created_at >= date_debut, StockMovement.created_at <= date_fin))
    elif date_debut:
        query = query.filter(StockMovement.created_at >= date_debut)
    elif date_fin:
        query = query.filter(StockMovement.created_at <= date_fin)
    movements = query.order_by(StockMovement.created_at.desc()).all()
    return movements

@router.get("/movements/{product_id}")
def get_product_movements(
    product_id: int,
    date_debut: Optional[datetime] = Query(None),
    date_fin: Optional[datetime] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Historique des mouvements pour un produit"""
    
    query = db.query(StockMovement).filter(StockMovement.product_id == product_id)
    
    if date_debut and date_fin:
        query = query.filter(
            and_(StockMovement.created_at >= date_debut, StockMovement.created_at <= date_fin)
        )
    elif date_debut:
        query = query.filter(StockMovement.created_at >= date_debut)
    elif date_fin:
        query = query.filter(StockMovement.created_at <= date_fin)
    
    movements = query.order_by(StockMovement.created_at.desc()).all()
    return movements

@router.get("/low-stock")
def get_low_stock_alert(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Produits avec stock faible (en dessous du seuil d'alerte)"""
    
    products = db.query(Product).filter(
        (Product.stock_actuel_kg <= Product.seuil_alerte) |
        (Product.stock_actuel_cartons <= Product.seuil_alerte)
    ).all()
    
    return {
        "produits_alerte": len(products),
        "details": [
            {
                "produit": product.nom_produit,
                "code_produit": product.code_produit,
                "stock_kg": product.stock_actuel_kg,
                "stock_cartons": product.stock_actuel_cartons,
                "seuil_alerte": product.seuil_alerte
            }
            for product in products
        ]
    }

@router.get("/expired-products")
def get_expired_products(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Produits périmés ou qui vont expirer bientôt"""
    
    today = datetime.now()
    next_week = today + timedelta(days=7)
    
    # Entrées qui vont expirer bientôt
    expiring_entries = db.query(StockEntry).filter(
        and_(
            StockEntry.date_peremption.isnot(None),
            StockEntry.date_peremption <= next_week
        )
    ).all()
    
    return {
        "produits_expirant": len(expiring_entries),
        "details": [
            {
                "produit": entry.product.nom_produit,
                "code_produit": entry.product.code_produit,
                "date_peremption": entry.date_peremption,
                "qte_kg": entry.qte_kg,
                "qte_cartons": entry.qte_cartons,
                "num_reception": entry.num_reception
            }
            for entry in expiring_entries
        ]
    }

def create_pdf_report(content: dict, title: str) -> str:
    """Créer un rapport PDF"""
    
    # Créer un fichier temporaire
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    temp_filename = temp_file.name
    temp_file.close()
    
    # Créer le PDF
    c = canvas.Canvas(temp_filename, pagesize=A4)
    width, height = A4
    
    # Titre
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, height - 50, title)
    
    # Date de génération
    c.setFont("Helvetica", 12)
    c.drawString(50, height - 80, f"Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    # Contenu (exemple simplifié)
    y_position = height - 120
    c.setFont("Helvetica", 10)
    
    for key, value in content.items():
        if y_position < 50:  # Nouvelle page si nécessaire
            c.showPage()
            y_position = height - 50
        
        c.drawString(50, y_position, f"{key}: {value}")
        y_position -= 20
    
    c.save()
    return temp_filename

@router.get("/pdf/stock-summary")
def download_stock_summary_pdf(
    date_debut: Optional[datetime] = Query(None),
    date_fin: Optional[datetime] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Télécharger le résumé du stock en PDF"""
    
    # Obtenir les données
    stock_data = get_stock_summary(date_debut, date_fin, db, current_user)
    
    # Préparer le contenu pour le PDF
    content = {}
    for i, item in enumerate(stock_data):
        content[f"Produit {i+1}"] = f"{item.product.nom_produit} - Stock: {item.stock_actuel_kg}kg, {item.stock_actuel_cartons} cartons"
    
    # Créer le PDF
    pdf_path = create_pdf_report(content, "Résumé du Stock")
    
    return FileResponse(
        pdf_path,
        media_type='application/pdf',
        filename=f"resume_stock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    )

@router.get("/excel/stock-summary")
def download_stock_summary_excel(
    date_debut: Optional[datetime] = Query(None),
    date_fin: Optional[datetime] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Télécharger le résumé du stock en Excel"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        # Obtenir les données
        stock_data = get_stock_summary(date_debut, date_fin, db, current_user)
        
        # Créer le classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Résumé Stock"
        
        # En-têtes
        headers = ["Code Produit", "Nom Produit", "Stock KG", "Stock Cartons", "Entrées KG", "Entrées Cartons", "Sorties KG", "Sorties Cartons"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for row, item in enumerate(stock_data, 2):
            ws.cell(row=row, column=1, value=item.product.code_produit)
            ws.cell(row=row, column=2, value=item.product.nom_produit)
            ws.cell(row=row, column=3, value=item.stock_actuel_kg)
            ws.cell(row=row, column=4, value=item.stock_actuel_cartons)
            ws.cell(row=row, column=5, value=item.total_entrees_kg)
            ws.cell(row=row, column=6, value=item.total_entrees_cartons)
            ws.cell(row=row, column=7, value=item.total_sorties_kg)
            ws.cell(row=row, column=8, value=item.total_sorties_cartons)
        
        # Sauvegarder dans un fichier temporaire
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return FileResponse(
            temp_file.name,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f"resume_stock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not available for Excel export")

@router.get("/pdf/stock-reception")
def download_stock_reception_pdf(
    num_reception: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Télécharger le bon d'entrée complet (réception) en PDF avec toutes les lignes (items)."""
    entries = db.query(StockEntry).filter(StockEntry.num_reception == num_reception).all()
    if not entries:
        raise HTTPException(status_code=404, detail="Aucune entrée trouvée pour ce numéro de réception")

    # Préparer en-tête
    head = entries[0]
    header = {
        "Numéro de Réception": head.num_reception,
        "Date de Réception": head.date_reception.strftime('%d/%m/%Y %H:%M'),
        "Numéro Carnet": head.num_reception_carnet or '-',
        "Numéro Facture": head.num_facture or '-',
        "Numéro Packing Liste": head.num_packing_liste or '-',
        "Nombre d'articles": len(entries)
    }

    # Créer un fichier temporaire PDF avec tableau simple
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdf_path = temp_file.name
    temp_file.close()

    c = canvas.Canvas(pdf_path, pagesize=A4)
    width, height = A4

    # Titre
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, height - 50, f"Bon d'entrée - Réception {num_reception}")

    # En-tête
    c.setFont("Helvetica", 10)
    y = height - 80
    for k, v in header.items():
        c.drawString(50, y, f"{k}: {v}")
        y -= 16

    y -= 10
    # Entêtes du tableau des items
    c.setFont("Helvetica-Bold", 10)
    headers = ["Code Produit", "Nom Produit", "Quantité (kg)", "Quantité (cartons)", "Date Péremption", "Remarques"]
    col_x = [50, 160, 350, 450, 540, 630]
    for i, htxt in enumerate(headers):
        c.drawString(col_x[i], y, htxt)
    y -= 14
    c.setFont("Helvetica", 9)

    # Lignes des items
    for e in entries:
        if y < 60:
            c.showPage()
            y = height - 50
            c.setFont("Helvetica-Bold", 10)
            for i, htxt in enumerate(headers):
                c.drawString(col_x[i], y, htxt)
            y -= 14
            c.setFont("Helvetica", 9)
        date_per = e.date_peremption.strftime('%d/%m/%Y') if e.date_peremption else '-'
        c.drawString(col_x[0], y, f"{e.product.code_produit}")
        c.drawString(col_x[1], y, f"{e.product.nom_produit}")
        c.drawString(col_x[2], y, f"{e.qte_kg}")
        c.drawString(col_x[3], y, f"{e.qte_cartons}")
        c.drawString(col_x[4], y, f"{date_per}")
        c.drawString(col_x[5], y, f"{e.remarque or ''}")
        y -= 14

    c.save()

    return FileResponse(
        pdf_path,
        media_type='application/pdf',
        filename=f"bon_entree_{num_reception}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    )

@router.get("/export-data")
def export_data(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Exporter les données de stock en JSON."""
    try:
        # Récupérer les produits et leurs stocks
        products = db.query(Product).all()
        data = []
        for product in products:
            data.append({
                "code_produit": product.code_produit,
                "nom_produit": product.nom_produit,
                "stock_actuel_kg": product.stock_actuel_kg,
                "stock_actuel_cartons": product.stock_actuel_cartons,
                "prix_achat": product.prix_achat,
                "seuil_alerte": product.seuil_alerte
            })
        return {"status": "success", "data": data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'exportation des données: {str(e)}")
